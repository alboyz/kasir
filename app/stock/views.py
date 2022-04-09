from django.shortcuts import render, redirect
from django.http import HttpResponse
from datetime import datetime
from django.views import generic
from .models import Stock
import csv
from openpyxl import Workbook
from django.contrib import messages
from .forms import StockCreateForm, StockSearchForm, StockUpdateForm
# Create your views here.


class IndexPageView(generic.TemplateView):
    template_name = "dashboard/index.html"

    def get_context_data(self, *args, **kwargs):
        context = super(IndexPageView,  self).get_context_data(
            *args, **kwargs)
        context['title'] = "Welcome This Is Home Page"
        return context


def list_item(request):
    title = "List Item"
    form = StockSearchForm(request.POST or None)
    queryset = Stock.objects.all()
    context = {
        "title": title,
        "queryset": queryset,
        "form": form,
    }
    if request.method == 'POST':

        '''Search'''
        queryset = Stock.objects.filter(
            catagory__icontains=form['catagory'].value(),
            item_name__icontains=form['item_name'].value()
        )

        context = {
            "form": form,
            "title": title,
            "queryset": queryset,
        }
    return render(request, "dashboard/list_item.html", context)


def add_item(request):
    form = StockCreateForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Berhasil Disimpan")
        return redirect('/list_item')
    context = {
        "form": form,
        "title": "Add Item",
    }
    return render(request, "dashboard/add_item.html", context)


def update_item(request, pk):
    queryset = Stock.objects.get(id=pk)
    form = StockUpdateForm(instance=queryset)
    if request.method == 'POST':
        form = StockUpdateForm(request.POST, instance=queryset)
        if form.is_valid():
            form.save()
            return redirect('/list_item')

    context = {
        "form": form
    }

    return render(request, 'dashboard/add_item.html', context)


def delete_item(request, pk):
    queryset = Stock.objects.get(id=pk)
    form = StockUpdateForm(instance=queryset)
    if request.method == "POST":
        queryset.delete()
        messages.success(request, "Berhasil Hapus")

        return redirect("/list_item")
    return render(request, "dashboard/confirm_delete.html")


def exportCSV(request):
    queryset = Stock.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment filename="{}.csv"' +\
        str(datetime.now)
    writer = csv.writer(response)
    writer.writerow(['CATAGORY', 'ITEM NAME', 'QUANTITY'])
    list_fields = queryset.values_list('catagory', 'item_name', 'quantity')
    for queryset in list_fields:
        writer.writerow(queryset)
    return response


def exportEXCEL(request):
    queryset = Stock.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=list-item.xlsx'.format(
        datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'List Item'

    # Define the titles for columns
    columns = [
        'CATAGORY', 'ITEM NAME', 'QUANTITY'
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all list
    for instance in queryset:
        row_num += 1
        row = [
            instance.catagory,
            instance.item_name,
            instance.quantity,
        ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
